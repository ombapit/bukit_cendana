import re
import psycopg2
import openpyxl

XLS_PATH = r"C:\data_dave\work\alone\david\bukitcendana\aplikasi\doc\Database_Warga_Bukit_Cendana_Qurban.xlsx"

DB_CONFIG = {
    "host": "157.230.244.171",
    "port": 5433,
    "database": "bukitcendana",
    "user": "shareliveloc",
    "password": "Bismillah123",
}

SKIP_SHEETS = {"Data Lama", "Data Warga_Sofie", "Lansia", "Update (Gabungan)"}

KK_KEYWORDS = {"KEPALA"}

def is_kk(status: str) -> bool:
    s = status.upper()
    return any(k in s for k in KK_KEYWORDS)

def is_valid_blok(v) -> bool:
    return bool(v and str(v).strip() and str(v).strip() not in ("-", "None"))

def clean(v) -> str:
    if not v:
        return ""
    s = str(v).strip()
    return "" if s in ("-", "None", "nan") else s

def read_xls():
    """
    Returns list of dicts: {blok, nama, status_hubungan, no_telp}
    Only non-KK rows with a valid blok and nama.
    """
    wb = openpyxl.load_workbook(XLS_PATH, data_only=True)
    anggota = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row
        header_row_idx = None
        for i, row in enumerate(rows):
            vals = [str(v).upper().strip() if v else "" for v in row]
            if "BLOK" in vals and any("NAMA" in v for v in vals):
                header_row_idx = i
                break
        if header_row_idx is None:
            continue

        headers = [str(v).upper().strip() if v else "" for v in rows[header_row_idx]]

        col_blok   = next((i for i, h in enumerate(headers) if h == "BLOK"), None)
        col_nama   = next((i for i, h in enumerate(headers) if "NAMA" in h), None)
        col_wa     = next((i for i, h in enumerate(headers) if "WHATSAP" in h or ("WA" in h and "NO" in h)), None)
        if col_wa is None:
            col_wa = next((i for i, h in enumerate(headers) if "WA" in h), None)
        col_status = next((i for i, h in enumerate(headers) if "STATUS" in h and ("HUB" in h or "KELUARGA" in h)), None)

        if col_blok is None or col_nama is None:
            continue

        current_blok = None

        for row in rows[header_row_idx + 1:]:
            blok_val   = row[col_blok]   if col_blok < len(row)                          else None
            nama_val   = row[col_nama]   if col_nama < len(row)                          else None
            wa_val     = row[col_wa]     if col_wa is not None and col_wa < len(row)     else None
            status_val = row[col_status] if col_status is not None and col_status < len(row) else None

            # Update current blok when row has a new valid blok
            if is_valid_blok(blok_val) and re.match(r'^[A-Za-z]+\d+/\d+$', str(blok_val).strip()):
                current_blok = str(blok_val).strip()

            if not current_blok:
                continue

            nama_str   = clean(nama_val)
            status_str = clean(status_val)
            wa_str     = clean(wa_val)

            if not nama_str:
                continue

            # Skip Kepala Keluarga — already in warga table
            if is_kk(status_str):
                continue

            # Skip rows with no meaningful status and no phone (likely filler rows)
            if not status_str and not wa_str and not nama_str:
                continue

            anggota.append({
                "blok":            current_blok,
                "nama":            nama_str,
                "status_hubungan": status_str,
                "no_telp":         wa_str,
            })

    return anggota

def main():
    print("Reading XLS...")
    anggota = read_xls()
    print(f"Non-KK anggota from XLS: {len(anggota)}")

    print("\nConnecting to DB...")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Build blok -> warga_id map from DB
    cur.execute("SELECT id, blok FROM warga")
    warga_rows = cur.fetchall()
    warga_by_blok = {row[1]: str(row[0]) for row in warga_rows}
    print(f"Warga in DB: {len(warga_rows)}")

    # Truncate existing anggota_keluarga (idempotent seed)
    cur.execute("TRUNCATE TABLE anggota_keluarga RESTART IDENTITY CASCADE")
    print("Truncated anggota_keluarga")

    inserted = 0
    skipped_no_warga = []

    for rec in anggota:
        warga_id = warga_by_blok.get(rec["blok"])
        if not warga_id:
            skipped_no_warga.append(rec)
            continue
        cur.execute(
            """
            INSERT INTO anggota_keluarga (warga_id, nama, status_hubungan, no_telp)
            VALUES (%s, %s, %s, %s)
            """,
            (warga_id, rec["nama"], rec["status_hubungan"], rec["no_telp"])
        )
        inserted += 1

    conn.commit()
    print(f"Inserted {inserted} anggota_keluarga records.")

    if skipped_no_warga:
        print(f"\nSkipped {len(skipped_no_warga)} rows (blok not found in warga table):")
        for r in skipped_no_warga[:20]:
            print(f"  blok={r['blok']}  nama={r['nama']}  status={r['status_hubungan']}")
        if len(skipped_no_warga) > 20:
            print(f"  ... and {len(skipped_no_warga) - 20} more")

    cur.close()
    conn.close()
    print("Done.")

if __name__ == "__main__":
    main()
