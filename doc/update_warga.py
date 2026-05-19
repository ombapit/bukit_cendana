import re
import psycopg2
import openpyxl

XLS_PATH = r"C:\data_dave\work\alone\david\bukitcendana\aplikasi\doc\Database_Warga_Bukit_Cendana_Qurban.xlsx"

DB_CONFIG = {
    "host": "157.230.244.171",
    "port": 5433,
    "database": "bukitcendana",
    "user": "shareliveloc",
    "password": "Bismillah123",
}

SKIP_SHEETS = {"Data Lama", "Data Warga_Sofie", "Lansia", "Update (Gabungan)"}

def is_valid_blok(v):
    return bool(v and str(v).strip() and str(v).strip() != '-' and str(v).strip() != 'None')

def clean_phone(v):
    if not v:
        return ""
    s = str(v).strip()
    if s in ("-", "None", "nan", ""):
        return ""
    return s

def read_xls():
    wb = openpyxl.load_workbook(XLS_PATH, data_only=True)
    records = {}  # blok_xls -> {nama, no_telp}

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # Find header row: contains 'BLOK' and 'NAMA'
        header_row_idx = None
        for i, row in enumerate(rows):
            vals = [str(v).upper().strip() if v else "" for v in row]
            if "BLOK" in vals and any("NAMA" in v for v in vals):
                header_row_idx = i
                break
        if header_row_idx is None:
            print(f"  [SKIP] {sheet_name}: no header")
            continue

        headers = [str(v).upper().strip() if v else "" for v in rows[header_row_idx]]

        # Find columns by header name
        col_blok = next((i for i, h in enumerate(headers) if h == "BLOK"), None)
        col_nama = next((i for i, h in enumerate(headers) if "NAMA" in h), None)
        col_wa   = next((i for i, h in enumerate(headers) if "WHATSAP" in h or "WA" in h or "TELP" in h or "HP" in h), None)
        col_status = next((i for i, h in enumerate(headers) if "STATUS" in h and ("HUB" in h or "KEL" in h or "KELUARGA" in h)), None)

        if col_blok is None or col_nama is None:
            print(f"  [SKIP] {sheet_name}: missing BLOK/NAMA col")
            continue

        current_blok = None

        # Data rows start after header (skip number row if any)
        for row in rows[header_row_idx + 1:]:
            blok_val = row[col_blok] if col_blok < len(row) else None
            nama_val = row[col_nama] if col_nama < len(row) else None
            wa_val   = row[col_wa]   if col_wa is not None and col_wa < len(row) else None
            status_val = row[col_status] if col_status is not None and col_status < len(row) else None

            # Track current blok
            if is_valid_blok(blok_val) and re.match(r'^[A-Za-z]+\d+/\d+$', str(blok_val).strip()):
                current_blok = str(blok_val).strip()

            if not current_blok:
                continue
            if not nama_val or str(nama_val).strip() in ("-", "None", ""):
                continue

            nama_str   = str(nama_val).strip()
            wa_str     = clean_phone(wa_val)
            status_str = str(status_val).strip().upper() if status_val else ""

            is_kk = "KEPALA" in status_str

            if current_blok not in records:
                # First row for this blok — store it
                records[current_blok] = {"nama": nama_str, "no_telp": wa_str}
            elif is_kk:
                # Override with KK row
                records[current_blok] = {"nama": nama_str, "no_telp": wa_str}

    return records

def xls_blok_to_db(xls_blok: str) -> str:
    """J00/01 -> J.00/ 01  (also handles J20/2 -> J.20/ 02 via zero-pad)"""
    m = re.match(r'^([A-Za-z]+)(\d+)/(\d+)$', xls_blok.strip())
    if not m:
        return xls_blok
    prefix, area, unit = m.group(1), m.group(2), m.group(3)
    # Zero-pad unit to match DB format (e.g. '2' -> '02')
    if len(unit) == 1:
        unit = unit.zfill(2)
    return f"{prefix}.{area}/ {unit}"

def main():
    print("Reading XLS...")
    xls_data = read_xls()
    print(f"Found {len(xls_data)} unique blok entries in XLS")

    print("\nConnecting to DB...")
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    cur.execute("SELECT id, nama, blok, no_telp FROM warga ORDER BY blok")
    db_rows = cur.fetchall()
    db_by_blok = {row[2]: {"id": row[0], "nama": row[1], "no_telp": row[3]} for row in db_rows}
    print(f"DB has {len(db_rows)} warga records")

    matched = []
    not_in_db = []
    all_db_blok_xls_equiv = {xls_blok_to_db(b) for b in xls_data} | set(xls_data.keys())
    not_in_xls = [b for b in db_by_blok if b not in all_db_blok_xls_equiv]

    for xls_blok, xls_rec in sorted(xls_data.items()):
        # Try direct match first (already updated bloks), then old DB format
        if xls_blok in db_by_blok:
            matched.append((xls_blok, xls_blok, xls_rec, db_by_blok[xls_blok]))
        else:
            db_blok = xls_blok_to_db(xls_blok)
            if db_blok in db_by_blok:
                matched.append((xls_blok, db_blok, xls_rec, db_by_blok[db_blok]))
            else:
                not_in_db.append((xls_blok, db_blok, xls_rec))

    print(f"\nMatched: {len(matched)}")
    print(f"XLS bloks NOT in DB: {len(not_in_db)}")
    print(f"DB bloks NOT in XLS: {len(not_in_xls)}")

    if not_in_db:
        print("\n=== XLS bloks NOT found in DB ===")
        for xls_blok, db_blok, rec in not_in_db:
            print(f"  XLS={xls_blok} -> expected DB={repr(db_blok)}  nama={rec['nama']}")

    if not_in_xls:
        print("\n=== DB bloks NOT in XLS ===")
        for db_blok in sorted(not_in_xls):
            rec = db_by_blok[db_blok]
            print(f"  DB blok={repr(db_blok)}  nama={rec['nama']}")

    # UPDATE: nama, blok (XLS format J00/01), no_telp — id unchanged
    print(f"\n=== Updating {len(matched)} warga records ===")
    for xls_blok, db_blok, xls_rec, db_rec in matched:
        cur.execute(
            "UPDATE warga SET nama=%s, blok=%s, no_telp=%s WHERE id=%s",
            (xls_rec["nama"], xls_blok, xls_rec["no_telp"], db_rec["id"])
        )

    conn.commit()
    print(f"Updated {len(matched)} records (nama, blok=XLS format, no_telp). id unchanged.")
    cur.close()
    conn.close()
    print("Done.")

if __name__ == "__main__":
    main()
